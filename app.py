import streamlit as st
import pandas as pd
import numpy as np
import random
import math
import io
from copy import deepcopy
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="Formador de Grupos",
    page_icon="👥",
    layout="wide",
)

# ============================================================
# CSS PERSONALIZADO
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Serif+Display&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.stApp { background-color: #F7F5F2; }

#MainMenu, footer, header { visibility: hidden; }

.main-header {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 60%, #0f3460 100%);
    border-radius: 20px;
    padding: 40px 48px;
    margin-bottom: 32px;
    color: white;
    position: relative;
    overflow: hidden;
}
.main-header::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(229,160,80,0.25) 0%, transparent 70%);
    border-radius: 50%;
}
.main-header h1 {
    font-family: 'DM Serif Display', serif;
    font-size: 2.4rem;
    font-weight: 400;
    margin: 0 0 8px 0;
}
.main-header p {
    font-size: 1rem;
    color: rgba(255,255,255,0.65);
    margin: 0;
    font-weight: 300;
}

.section-card {
    background: white;
    border-radius: 16px;
    padding: 28px 32px;
    margin-bottom: 24px;
    border: 1px solid rgba(0,0,0,0.06);
    box-shadow: 0 2px 12px rgba(0,0,0,0.05);
}
.step-badge {
    display: inline-block;
    background: #1a1a2e;
    color: white;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 2px;
    text-transform: uppercase;
    padding: 4px 12px;
    border-radius: 20px;
    margin-bottom: 10px;
}
.section-title {
    font-family: 'DM Serif Display', serif;
    font-size: 1.5rem;
    font-weight: 400;
    color: #1a1a2e;
    margin: 0 0 4px 0;
}
.section-subtitle {
    color: #888;
    font-size: 0.9rem;
    margin: 0 0 24px 0;
}

.familia-card {
    background: white;
    border-radius: 14px;
    padding: 20px;
    border: 1px solid rgba(0,0,0,0.07);
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    margin-bottom: 16px;
}
.familia-titulo {
    font-family: 'DM Serif Display', serif;
    font-size: 1.15rem;
    color: #1a1a2e;
    margin: 0 0 6px 0;
}
.familia-meta {
    font-size: 0.8rem;
    color: #999;
    margin: 0 0 14px 0;
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
}
.meta-chip {
    background: #F0EDE8;
    border-radius: 20px;
    padding: 2px 10px;
    color: #555;
    font-weight: 500;
}

.personas-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
.personas-table th {
    text-align: left;
    color: #aaa;
    font-weight: 600;
    font-size: 0.72rem;
    letter-spacing: 1px;
    text-transform: uppercase;
    padding: 0 8px 8px 8px;
    border-bottom: 1px solid #eee;
}
.personas-table td {
    padding: 7px 8px;
    color: #333;
    border-bottom: 1px solid #f5f5f5;
    vertical-align: middle;
}
.personas-table tr:last-child td { border-bottom: none; }

.badge-hombre {
    background: #EBF5FB; color: #2980B9;
    border-radius: 20px; padding: 2px 9px;
    font-size: 0.75rem; font-weight: 600;
}
.badge-mujer {
    background: #FDEDEC; color: #C0392B;
    border-radius: 20px; padding: 2px 9px;
    font-size: 0.75rem; font-weight: 600;
}
.lider-badge {
    background: #FEF9E7; color: #D4AC0D;
    border-radius: 20px; padding: 1px 8px;
    font-size: 0.72rem; font-weight: 600;
    margin-left: 6px;
}

.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1a1a2e, #0f3460) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: 12px 0 !important;
}

.stTabs [data-baseweb="tab-list"] {
    background: #F0EDE8;
    border-radius: 12px;
    padding: 4px;
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 9px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    color: #888 !important;
}
.stTabs [aria-selected="true"] {
    background: white !important;
    color: #1a1a2e !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.1) !important;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# HEADER
# ============================================================
st.markdown("""
<div class="main-header">
    <h1>👥 Formador de Grupos</h1>
    <p>Genera grupos balanceados automáticamente considerando edad, carrera, género y restricciones personalizadas.</p>
</div>
""", unsafe_allow_html=True)

# ============================================================
# FUNCIONES DE OPTIMIZACIÓN
# ============================================================
def calcular_score(grupos, edades, es_hombre, carreras, pares_idx, piso_h, techo_h):
    score = 0
    varianzas = []
    for g_indices in grupos:
        edades_g   = edades[g_indices]
        hombres_g  = es_hombre[g_indices].sum()
        carreras_g = carreras[g_indices]
        idx_set    = set(g_indices)
        for idx1, idx2 in pares_idx:
            if idx1 in idx_set and idx2 in idx_set:
                score += 1_000_000
        for conteo in Counter(carreras_g).values():
            if conteo > 1:
                score += (conteo - 1) * 5_000
        if hombres_g < piso_h or hombres_g > techo_h:
            score += 10_000
        v = edades_g.var()
        varianzas.append(v)
        score += v * 150
    score += max(varianzas) * 600
    score += (max(varianzas) - min(varianzas)) * 300
    return score


def inicializar_grupos(df, lideres, n_grupos):
    idx_lideres = df[df["Nombre"].isin(lideres)].index.tolist()
    idx_resto   = df[~df["Nombre"].isin(lideres)].index.tolist()
    random.shuffle(idx_resto)
    grupos = [[] for _ in range(n_grupos)]
    for i, idx in enumerate(idx_lideres):
        grupos[i].append(idx)
    ptr = 0
    for g_idx in range(n_grupos):
        while len(grupos[g_idx]) < len(df) // n_grupos:
            grupos[g_idx].append(idx_resto[ptr])
            ptr += 1
    while ptr < len(idx_resto):
        for g_idx in range(n_grupos):
            if ptr < len(idx_resto):
                grupos[g_idx].append(idx_resto[ptr])
                ptr += 1
    return [list(g) for g in grupos]


def optimizar(df, lideres, pares_prohibidos, n_grupos, n_iter=80_000, T_inicial=500.0, T_final=0.1):
    edades    = df["Edad"].values.astype(float)
    es_hombre = (df["Sexo"] == "Hombre").values
    carreras  = df["Carrera"].values
    n_hombres_total = es_hombre.sum()
    piso_h  = math.floor(n_hombres_total / n_grupos)
    techo_h = math.ceil(n_hombres_total / n_grupos)
    pares_idx = []
    for p1, p2 in pares_prohibidos:
        r1 = df[df["Nombre"] == p1]
        r2 = df[df["Nombre"] == p2]
        if not r1.empty and not r2.empty:
            pares_idx.append((r1.index[0], r2.index[0]))
    grupos = inicializar_grupos(df, lideres, n_grupos)
    score_actual = calcular_score(grupos, edades, es_hombre, carreras, pares_idx, piso_h, techo_h)
    mejor_grupos = deepcopy(grupos)
    mejor_score  = score_actual
    alpha = (T_final / T_inicial) ** (1.0 / n_iter)
    T = T_inicial
    n_lideres = len(lideres)
    for _ in range(n_iter):
        g1, g2 = random.sample(range(n_grupos), 2)
        start1 = 1 if g1 < n_lideres else 0
        start2 = 1 if g2 < n_lideres else 0
        p1 = random.randint(start1, len(grupos[g1]) - 1)
        p2 = random.randint(start2, len(grupos[g2]) - 1)
        grupos[g1][p1], grupos[g2][p2] = grupos[g2][p2], grupos[g1][p1]
        nuevo_score = calcular_score(grupos, edades, es_hombre, carreras, pares_idx, piso_h, techo_h)
        delta = nuevo_score - score_actual
        if delta < 0 or random.random() < math.exp(-delta / T):
            score_actual = nuevo_score
            if score_actual < mejor_score:
                mejor_score  = score_actual
                mejor_grupos = deepcopy(grupos)
        else:
            grupos[g1][p1], grupos[g2][p2] = grupos[g2][p2], grupos[g1][p1]
        T *= alpha
    return mejor_grupos, mejor_score


def solucion_a_frozenset(grupos):
    return frozenset(frozenset(g) for g in grupos)


MAX_OPCIONES = 5

def correr_optimizacion(df, lideres, pares_prohibidos, n_grupos, n_corridas=10):
    mejores = []
    huellas = set()
    mejor_score_global = float("inf")
    bar = st.progress(0, text="Iniciando optimización...")
    for semilla in range(n_corridas):
        bar.progress((semilla + 1) / n_corridas,
                     text=f"Analizando configuración {semilla + 1} de {n_corridas}...")
        random.seed(semilla)
        grupos, score = optimizar(df, lideres, pares_prohibidos, n_grupos)
        huella = solucion_a_frozenset(grupos)
        if score < mejor_score_global - 0.01:
            mejor_score_global = score
            mejores = [grupos]
            huellas = {huella}
        elif abs(score - mejor_score_global) < 0.01 and huella not in huellas:
            if len(mejores) < MAX_OPCIONES:
                mejores.append(grupos)
                huellas.add(huella)
    bar.empty()
    return mejores, mejor_score_global


# ============================================================
# FUNCIÓN: EXCEL DE RESULTADOS
# ============================================================
def generar_excel_resultados(df, mejores_resultados, lideres):
    wb = Workbook()
    wb.remove(wb.active)
    AZUL    = "1a1a2e"
    AZUL_CL = "EAE8F0"
    COLORES = ["E8F5E9","FFF3E0","FCE4EC","E3F2FD","F3E5F5",
               "E0F7FA","FBE9E7","E8EAF6","E0F2F1","FFF8E1"]
    borde = Border(
        left=Side(style="thin", color="DEDEDE"),
        right=Side(style="thin", color="DEDEDE"),
        top=Side(style="thin", color="DEDEDE"),
        bottom=Side(style="thin", color="DEDEDE"),
    )
    for n_op, grupos in enumerate(mejores_resultados):
        ws = wb.create_sheet(title=f"Opción {n_op + 1}")
        fila = 1
        for i, g_indices in enumerate(grupos):
            g    = df.loc[g_indices]
            h    = (g["Sexo"] == "Hombre").sum()
            m    = (g["Sexo"] == "Mujer").sum()
            prom = g["Edad"].mean()
            vari = g["Edad"].var()
            c    = COLORES[i % len(COLORES)]
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=5)
            t = ws.cell(row=fila, column=1,
                        value=f"  FAMILIA {i+1}   |   {h}H / {m}M   |   Edad promedio: {prom:.1f}   |   Varianza: {vari:.2f}")
            t.font      = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
            t.fill      = PatternFill("solid", start_color=AZUL)
            t.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[fila].height = 22
            fila += 1
            for col, hdr in enumerate(["Nombre","Sexo","Edad","Carrera","Rol"], 1):
                cell = ws.cell(row=fila, column=col, value=hdr)
                cell.font      = Font(bold=True, name="Calibri", size=10, color="444444")
                cell.fill      = PatternFill("solid", start_color=AZUL_CL)
                cell.alignment = Alignment(horizontal="center")
                cell.border    = borde
            fila += 1
            for _, persona in g.iterrows():
                rol = "⭐ Líder" if persona["Nombre"] in lideres else "Miembro"
                for col, val in enumerate(
                    [persona["Nombre"], persona["Sexo"], persona["Edad"], persona["Carrera"], rol], 1
                ):
                    cell = ws.cell(row=fila, column=col, value=val)
                    cell.font      = Font(name="Calibri", size=10)
                    cell.fill      = PatternFill("solid", start_color=c)
                    cell.alignment = Alignment(horizontal="left" if col not in (2,3,5) else "center")
                    cell.border    = borde
                fila += 1
            fila += 1
        ws.column_dimensions["A"].width = 44
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 32
        ws.column_dimensions["E"].width = 12
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# PASO 1: PARTICIPANTES
# ============================================================
st.markdown("""
<div class="section-card">
    <div class="step-badge">Paso 1</div>
    <h2 class="section-title">Participantes</h2>
    <p class="section-subtitle">Ingresa la lista de personas que formarán los grupos.</p>
""", unsafe_allow_html=True)

modo = st.radio("Método", ["📂 Subir archivo Excel", "✏️ Captura manual"],
                horizontal=True, label_visibility="collapsed")

df_participantes = None

if modo == "📂 Subir archivo Excel":
    col_dl, col_up = st.columns([1, 2])
    with col_dl:
        try:
            with open("plantilla_participantes.xlsx", "rb") as f:
                st.download_button("⬇️ Descargar plantilla", data=f,
                                   file_name="plantilla_participantes.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except FileNotFoundError:
            pass
    with col_up:
        archivo = st.file_uploader("Sube tu Excel", type=["xlsx"], label_visibility="collapsed")

    if archivo:
        try:
            df_leido = pd.read_excel(archivo)
            cols_req = {"Nombre", "Sexo", "Edad", "Carrera"}
            if not cols_req.issubset(df_leido.columns):
                st.error(f"El archivo debe tener estas columnas: {cols_req}")
            elif not df_leido[df_leido["Sexo"].notna() & ~df_leido["Sexo"].isin(["Hombre","Mujer"])].empty:
                st.error("La columna 'Sexo' solo acepta exactamente 'Hombre' o 'Mujer'.")
            elif df_leido["Nombre"].duplicated().any():
                dupes = df_leido[df_leido["Nombre"].duplicated(keep=False)]["Nombre"].unique()
                st.error(f"Nombres duplicados: {list(dupes)}")
            else:
                df_participantes = df_leido[["Nombre","Sexo","Edad","Carrera"]].copy()
                df_participantes["Edad"] = df_participantes["Edad"].astype(int)
                st.success(f"✅ {len(df_participantes)} participantes cargados.")
                st.dataframe(df_participantes, use_container_width=True, hide_index=True)
        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")

else:
    if "participantes" not in st.session_state:
        st.session_state.participantes = []
    with st.form("form_agregar", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns([3,1,1,2])
        nombre  = c1.text_input("Nombre completo")
        sexo    = c2.selectbox("Sexo", ["Mujer","Hombre"])
        edad    = c3.number_input("Edad", min_value=15, max_value=80, value=20, step=1)
        carrera = c4.text_input("Carrera")
        if st.form_submit_button("➕ Agregar", use_container_width=True):
            if not nombre.strip() or not carrera.strip():
                st.warning("El nombre y la carrera no pueden estar vacíos.")
            elif any(p["Nombre"] == nombre.strip() for p in st.session_state.participantes):
                st.warning("Ya existe un participante con ese nombre.")
            else:
                st.session_state.participantes.append({
                    "Nombre": nombre.strip(), "Sexo": sexo,
                    "Edad": int(edad), "Carrera": carrera.strip(),
                })
    if st.session_state.participantes:
        df_manual = pd.DataFrame(st.session_state.participantes)
        st.dataframe(df_manual, use_container_width=True, hide_index=True)
        ci, cb = st.columns([3,1])
        ci.caption(f"{len(df_manual)} participantes agregados.")
        if cb.button("🗑️ Borrar todos"):
            st.session_state.participantes = []
            st.rerun()
        df_participantes = df_manual.copy()

st.markdown("</div>", unsafe_allow_html=True)

# ============================================================
# PASO 2: CONFIGURACIÓN
# ============================================================
if df_participantes is not None and len(df_participantes) > 0:
    st.markdown("""
    <div class="section-card">
        <div class="step-badge">Paso 2</div>
        <h2 class="section-title">Configuración</h2>
        <p class="section-subtitle">Define cómo deben formarse los grupos y las restricciones especiales.</p>
    """, unsafe_allow_html=True)

    nombres_disponibles = sorted(df_participantes["Nombre"].tolist())
    col_l, col_r = st.columns(2)

    with col_l:
        n_grupos = st.number_input("¿Cuántos grupos?", min_value=2,
                                   max_value=len(df_participantes)//2,
                                   value=min(6, len(df_participantes)//5), step=1)
        st.markdown("**🏅 Líderes** — Personas que deben quedar en grupos distintos")
        lideres = st.multiselect("Líderes", options=nombres_disponibles,
                                 max_selections=n_grupos, label_visibility="collapsed",
                                 placeholder="Selecciona los líderes...")

    with col_r:
        st.markdown("**🚫 Pares prohibidos** — No pueden coincidir en el mismo grupo")
        if "pares" not in st.session_state:
            st.session_state.pares = []
        with st.form("form_par", clear_on_submit=True):
            pc1, pc2, pc3 = st.columns([2,2,1])
            par1 = pc1.selectbox("P1", nombres_disponibles, label_visibility="collapsed")
            par2 = pc2.selectbox("P2", nombres_disponibles, label_visibility="collapsed")
            if pc3.form_submit_button("➕", use_container_width=True):
                if par1 == par2:
                    st.warning("Selecciona dos personas distintas.")
                elif (par1,par2) in st.session_state.pares or (par2,par1) in st.session_state.pares:
                    st.warning("Ese par ya está registrado.")
                else:
                    st.session_state.pares.append((par1, par2))
        if st.session_state.pares:
            for idx, (p1, p2) in enumerate(st.session_state.pares):
                cp, cx = st.columns([5,1])
                cp.markdown(f"<small>🚫 {p1} &nbsp;↔&nbsp; {p2}</small>", unsafe_allow_html=True)
                if cx.button("✕", key=f"del_{idx}"):
                    st.session_state.pares.pop(idx)
                    st.rerun()
        else:
            st.caption("Sin pares prohibidos definidos.")

    st.markdown("</div>", unsafe_allow_html=True)

    # ============================================================
    # PASO 3: GENERAR
    # ============================================================
    st.markdown("""
    <div class="section-card">
        <div class="step-badge">Paso 3</div>
        <h2 class="section-title">Generar grupos</h2>
        <p class="section-subtitle">El algoritmo explorará múltiples configuraciones y presentará las mejores.</p>
    """, unsafe_allow_html=True)

    if len(df_participantes) % n_grupos != 0:
        t = len(df_participantes) // n_grupos
        st.warning(f"⚠️ {len(df_participantes)} participantes no se dividen exactamente en {int(n_grupos)} grupos. "
                   f"Algunos tendrán {t} personas y otros {t+1}.")

    if st.button("🚀  Generar grupos óptimos", type="primary", use_container_width=True):
        if len(df_participantes) < n_grupos:
            st.error("Hay menos participantes que grupos.")
        elif len(lideres) > n_grupos:
            st.error(f"Hay más líderes ({len(lideres)}) que grupos ({int(n_grupos)}).")
        else:
            mejores, score_global = correr_optimizacion(
                df_participantes, lideres, st.session_state.pares, int(n_grupos))
            st.session_state.mejores   = mejores
            st.session_state.score     = score_global
            st.session_state.df_result = df_participantes
            st.session_state.lideres   = lideres

    st.markdown("</div>", unsafe_allow_html=True)

    # ============================================================
    # PASO 4: RESULTADOS
    # ============================================================
    if "mejores" in st.session_state and st.session_state.mejores:
        mejores     = st.session_state.mejores
        df_res      = st.session_state.df_result
        lideres_res = st.session_state.get("lideres", [])

        st.markdown("""
        <div class="section-card">
            <div class="step-badge">Resultado</div>
            <h2 class="section-title">Grupos generados</h2>
        """, unsafe_allow_html=True)

        if len(mejores) > 1:
            st.info(f"Se encontraron **{len(mejores)} configuraciones distintas** con la misma calidad óptima. Elige la que prefieras.")

        tabs = st.tabs([f"Opción {i+1}" for i in range(len(mejores))]) if len(mejores) > 1 else [st.container()]

        for tab, grupos in zip(tabs, mejores):
            with tab:
                cols = st.columns(min(3, int(n_grupos)))
                for i, g_indices in enumerate(grupos):
                    g    = df_res.loc[g_indices]
                    h    = (g["Sexo"] == "Hombre").sum()
                    m    = (g["Sexo"] == "Mujer").sum()
                    prom = g["Edad"].mean()
                    vari = g["Edad"].var()
                    unic = g["Carrera"].nunique()

                    filas_html = ""
                    for _, p in g.iterrows():
                        badge = f'<span class="badge-{"hombre" if p["Sexo"]=="Hombre" else "mujer"}">{p["Sexo"]}</span>'
                        lider = f'<span class="lider-badge">⭐ Líder</span>' if p["Nombre"] in lideres_res else ""
                        filas_html += f"""<tr>
                            <td>{p['Nombre']}{lider}</td>
                            <td>{badge}</td>
                            <td style="color:#888">{int(p['Edad'])}</td>
                            <td style="color:#666">{p['Carrera']}</td>
                        </tr>"""

                    with cols[i % len(cols)]:
                        st.markdown(f"""
                        <div class="familia-card">
                            <p class="familia-titulo">Familia {i+1}</p>
                            <div class="familia-meta">
                                <span class="meta-chip">👥 {h}H / {m}M</span>
                                <span class="meta-chip">🎂 Prom. {prom:.1f}</span>
                                <span class="meta-chip">📊 Var. {vari:.2f}</span>
                                <span class="meta-chip">🎓 {unic}/{len(g)} carreras únicas</span>
                            </div>
                            <table class="personas-table">
                                <thead><tr>
                                    <th>Nombre</th><th>Sexo</th><th>Edad</th><th>Carrera</th>
                                </tr></thead>
                                <tbody>{filas_html}</tbody>
                            </table>
                        </div>""", unsafe_allow_html=True)

        st.markdown("</div>", unsafe_allow_html=True)

        excel_buf = generar_excel_resultados(df_res, mejores, lideres_res)
        st.download_button(
            "⬇️  Descargar resultados en Excel",
            data=excel_buf,
            file_name="grupos_resultado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )